from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import temp
import os
import shutil

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
 
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

curr_path = os.path.dirname(os.path.realpath(__file__))
shutil.copy(fr'{curr_path}\resources\invoice_template.xlsx', fr'{curr_path}\inprocess')

wb = load_workbook(filename='inprocess/invoice_template.xlsx')
sh = wb.active

def color_fill(color):
    for i in range(7):
        for cell in sh['A1:G1']:
            cell[i].fill = color
        for cell in sh['B18:F18']:
            try:
                cell[i].fill = color
            except IndexError:
                continue
        for cell in sh['A43:G43']:
            cell[i].fill = color

if temp.color=='Red':
    redFill = PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid')
    color_fill(redFill)

elif temp.color=='Green':
    greenFill = PatternFill(start_color='0099CC00',end_color='0099CC00',fill_type='solid')
    color_fill(greenFill)

elif temp.color=='Yellow':
    yellowFill = PatternFill(start_color='00FFCC00',end_color='00FFCC00',fill_type='solid')
    color_fill(yellowFill)

elif temp.color=='Blue':
    blueFill = PatternFill(start_color='000000FF',end_color='000000FF',fill_type='solid')
    color_fill(blueFill)

sh['B4'] = temp.bName
sh['B5'] = temp.bAddr
sh['B6'] = temp.bState
sh['B7'] = temp.bPhone
sh['B8'] = temp.bEmail

sh['F4'] = temp.date
sh['F6'] = temp.invoicenum

sh['B12'] = temp.btName
sh['B13'] = temp.btAddr
sh['B14'] = temp.btPhone
sh['B15'] = temp.btEmail

sh['D12'] = temp.stName
sh['D13'] = temp.stAddr
sh['D14'] = temp.stPhone

sh['F31'] = temp.Discount
sh['F33'] = temp.TaxRate
sh['F35'] = temp.Shipping

sh['B31'] = temp.Info

no_items = len(temp.items)
for i in range(1,no_items-11):
    sh.insert_rows(i+29)
    if i%2!=0:
        row29 = copyRange('A', '28', 'F', '28', sh)
        pasteRange('A', str(i+29), 'F', str(i+29), sh, row29)
    else:
        row30 = copyRange('A', '29', 'F', '29', sh)
        pasteRange('A', str(i+29), 'F', str(i+29), sh, row29)

for i in range(19,no_items+18):
    for j in temp.items:
        prop = j.split(' ')
        sh['B'+str(i)] = prop[0]
        sh['D'+str(i)] = prop[1]
        sh['E'+str(i)] = prop[2]

os.chdir(fr'{os.getcwd()}\inprocess')
wb.save(f'{temp.invoicenum}_{temp.btName}.xlsx')
os.chdir(os.path.dirname(os.getcwd()))
os.system('python -u "xlsxtopdf.py"')
